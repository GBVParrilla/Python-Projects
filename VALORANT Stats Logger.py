import openpyxl

path = "C:\\Users\\*insertusernamehere*\\Desktop\\valstats.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active
cell = sheet.cell(row = 1, column = 1)
print(cell.value)

while True:

    b = 1
    cell1 = sheet.cell(row = 1, column = b)
    if cell1.value != None:
        b = b + 1
    if cell1.value == None:
        print("")
    print(b)

    print("\n\nWELCOME TO VAL TRACKER\n")
    ans = input("Do you want to track or average?:")
    ans.lower()
    if ans == "track":
        data = [1,1,1,1,1,1,1,1,1,1,1,1]
        print("\nFILL UP STATS\n")
        data[0] = float(input("What is your combat score?"))
        data[1] = float(input("What is your KDA?"))
        data[2] = float(input("How many Kills?"))
        data[3] = float(input("How many Deaths?"))
        data[4] = float(input("How many Assists"))
        data[5] = float(input("What is your DMG per round?"))
        data[6] = float(input("How many First Bloods?"))
        data[7] = float(input("How many First to Dies?"))
        data[8] = float(input("How many Last to Dies?"))
        data[9] = float(input("What is your Headshot%?"))
        data[10] = float(input("What is your Bodyshots%?"))
        data[11] = float(input("What is your Legshot%?"))
        print("\n\n Match Comments:")
        if data[0] < 280:
            print("Your Combat Score is below average.")
        if data[1] < 1.7:
            print("Your KDA is below average.")
        if data[2] < 20:
            print("Your Kills are below average.")
        if data[3] > 10:
            print("Your Deaths are above average.")
        if data[4] < 2:
            print("Your Assists are below average.")
        if data[5] < 180:
            print("Your DMG per Round is below average.")
        if data[6] < 3:
            print("Your First Bloods are below average.")
        if data[7] > 2:
            print("Your First to Dies are above average.")
        if data[8] < 2:
            print("Your Last to Dies are below average.")
        if data[9] < 30:
            print("Your Headshot percentage is below average.")
        if data[10] > 68:
            print("Your Bodyshot percentage is above average.")
        if data[11] > 3:
            print("Your Legshot Percentage is above average.")
        sheet.append(data)
    if ans == "end":
        break
wb.save(path)